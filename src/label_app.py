#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import re
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook

KMART_TEMPLATE  = "KMART模板.xlsx"
TARGET_TEMPLATE = "TARGET模板.xlsx"
SUPPORTED_EXTS  = (".xlsx",)  # 只支持 xlsx（32bit 友好）


# ---------- 打包友好：读取内置资源 ----------
def resource_path(relpath: str) -> str:
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relpath)


# ---------- 工具函数 ----------
def choose_excel_file() -> str:
    root = tk.Tk()
    root.withdraw()
    try:
        file_path = filedialog.askopenfilename(
            title="请选择 Excel 文件（需含：销售合同/客户合同/客户简称/中文品名/产品编号/英文品名/合同数量/单箱 等列）",
            filetypes=[("Excel 文件", (".xlsx",)), ("所有文件", "*")],
        )
    finally:
        try:
            root.update()
        except Exception:
            pass
        root.destroy()

    if not file_path:
        print("❌ 未选择文件，程序结束。")
        sys.exit(0)

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_EXTS:
        messagebox.showerror("文件类型错误", f"请选择 .xlsx（当前：{ext}）")
        sys.exit(1)

    return os.path.normpath(file_path)


def is_na(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def s(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def fnum(v):
    try:
        if v is None:
            return None
        if isinstance(v, str):
            vv = v.replace(",", "").strip()
            if vv == "":
                return None
            return float(vv)
        return float(v)
    except Exception:
        return None


def extract_inside_brackets(text: str) -> str:
    if not text:
        return ""
    m = re.search(r"[（(]([^）)]+)[)）]", str(text))
    return m.group(1).strip() if m else ""


def before_bracket_digits(text: str) -> str:
    t = s(text)
    for br in ("（", "("):
        if br in t:
            t = t.split(br)[0]
            break
    return re.sub(r"[^0-9]", "", t)


def last_three_digits_padded(text: str) -> str:
    digits = "".join(re.findall(r"\d", s(text)))
    return digits[-3:].zfill(3) if digits else "000"


def append_keycode_to_a10(ws, code: str):
    if not code:
        return
    cell = ws["A10"]
    orig = "" if cell.value is None else str(cell.value)
    if "KEYCODE" not in orig.upper():
        cell.value = f"KEYCODE: {code}"
        return
    if code in orig:
        return
    if orig.rstrip().endswith((":", "：")):
        cell.value = orig + code
    else:
        cell.value = (orig + " " + code).strip()


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip()


def unique_path(dirpath: str, filename: str) -> str:
    base, ext = os.path.splitext(filename)
    cand = os.path.join(dirpath, filename)
    i = 2
    while os.path.exists(cand):
        cand = os.path.join(dirpath, f"{base}-{i}{ext}")
        i += 1
    return cand


def fmt_intlike(val) -> str:
    """整数样式：6/6.0/55.000/'21245862.0' -> '6'/'55'/'21245862'；否则原样。"""
    txt = s(val)
    if txt == "":
        return ""
    m = re.fullmatch(r"(\d+)\.0+", txt)
    if m:
        return m.group(1)
    try:
        f = float(txt.replace(",", ""))
        if abs(f - round(f)) < 1e-9:
            return str(int(round(f)))
    except Exception:
        pass
    return txt


def fmt_dim(val) -> str:
    """
    维度格式化（用于 E10 的 长/宽/高）：
    - 若数值是整数（21、21.0）→ '21'
    - 若有小数（21.5、21.50）→ 去除多余的尾随 0（'21.5'）
    - 无法解析数字则原样返回
    """
    if val is None:
        return ""
    raw = str(val).strip()
    if raw == "":
        return ""
    raw_clean = raw.replace(",", "")
    try:
        f = float(raw_clean)
    except Exception:
        return raw
    if abs(f - round(f)) < 1e-9:
        return str(int(round(f)))
    sflt = f"{f}"
    if "e" in sflt or "E" in sflt:
        sflt = f"{f:.10f}"
    sflt = sflt.rstrip("0").rstrip(".")
    return sflt


def read_xlsx_as_dict_rows(path: str):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    header = [("" if c is None else str(c).strip()) for c in rows[0]]
    data_rows = []
    for r in rows[1:]:
        d = {}
        for i, col in enumerate(header):
            if col == "":
                continue
            d[col] = r[i] if i < len(r) else None
        data_rows.append(d)
    return header, data_rows


def is_blank_row_dict(d: dict) -> bool:
    return all(is_na(v) for v in d.values())


# ---------- 主流程 ----------
def main():
    kmart_tpl  = resource_path(KMART_TEMPLATE)
    target_tpl = resource_path(TARGET_TEMPLATE)

    for p in (kmart_tpl, target_tpl):
        if not os.path.exists(p):
            messagebox.showerror(
                "模板缺失",
                f"未找到模板：{p}\n请确保把 {KMART_TEMPLATE} 和 {TARGET_TEMPLATE} 作为数据文件加入。"
            )
            sys.exit(1)

    in_path = choose_excel_file()
    out_dir = os.path.dirname(in_path)

    try:
        header, data_rows = read_xlsx_as_dict_rows(in_path)
    except Exception as e:
        messagebox.showerror("读取失败", f"无法读取 Excel：{e}")
        sys.exit(1)

    # ---- drop trailing blank rows and skip the final summary row ----
    last_idx = len(data_rows) - 1
    while last_idx >= 0 and is_blank_row_dict(data_rows[last_idx]):
        last_idx -= 1

    if last_idx >= 0:
        joined = "".join(str(v).strip() for v in data_rows[last_idx].values() if v is not None)
        if any(kw in joined for kw in ("汇总", "合计", "总计")):
            print(f"ℹ️ 检测到最后一行为汇总/合计（第 {last_idx+2} 行），已跳过该行生成。")
            data_rows = data_rows[:last_idx]

    need_cols = ["销售合同", "客户合同", "客户简称", "中文品名", "产品编号", "英文品名", "合同数量", "单箱"]
    miss_cols = [c for c in need_cols if c not in header]
    if miss_cols:
        messagebox.showerror("列缺失", f"输入文件缺少列：{miss_cols}")
        sys.exit(1)

    def get_height_raw(row: dict):
        if "高" in row:
            return row.get("高")
        if "髙" in row:
            return row.get("髙")
        return ""

    success, failed = 0, 0
    for idx, row in enumerate(data_rows):
        try:
            e_sales   = s(row.get("销售合同"))         # E
            f_order   = s(row.get("客户合同"))         # F
            g_client  = s(row.get("客户简称"))         # G
            h_cname   = s(row.get("中文品名"))         # H
            i_prod    = s(row.get("产品编号"))         # I
            j_ename   = s(row.get("英文品名"))         # J
            qty_total = fnum(row.get("合同数量"))       # 合同数量
            pcs_eachN = fnum(row.get("单箱"))          # 单箱

            gross_wt = s(row.get("毛重", ""))

            if qty_total is not None and pcs_eachN not in (None, 0):
                carton_count = round(qty_total / pcs_eachN, 2)
            else:
                carton_count = ""

            is_kmart_row = ("KMART" in g_client.upper())

            length_v = fmt_dim(row.get("长", ""))
            width_v  = fmt_dim(row.get("宽", ""))
            height_v = fmt_dim(get_height_raw(row))
            dept3    = last_three_digits_padded(g_client)

            if is_kmart_row:
                wb = load_workbook(kmart_tpl)
                ws = wb[wb.sheetnames[0]]

                a7_val = "NZ" if "NZ" in e_sales.upper() else "AU"
                ws["A7"].value = a7_val
                ws["A8"].value = f"DEPARTMENT NO.:{dept3}"

                ws["B9"].value = fmt_intlike(f_order)

                key_digits = before_bracket_digits(i_prod)
                ws["A10"].value = f"KEYCODE: {key_digits}" if key_digits else "KEYCODE:"
                ws["A11"].value = f"DESCRIPTION：{j_ename}" if j_ename else "DESCRIPTION："

                ws["A12"].value = f"QTY ISSUE PACK: {fmt_intlike(pcs_eachN)} pcs Only"
                ws["A13"].value = f"QTY SHIPPER PACK: {fmt_intlike(pcs_eachN)} pcs Only"

                ws["E9"].value = f"GRS.WT.: {gross_wt} KGS"
                ws["E10"].value = f"D:{length_v}×{width_v}×{height_v}CMS"

                ws["C14"].value = carton_count

                out_name = f"{e_sales} {key_digits}#{h_cname} 唛头（{a7_val}）.xlsx"

            else:
                wb = load_workbook(target_tpl)
                ws = wb[wb.sheetnames[0]]

                ws["A8"].value = f"DEPARTMENT NO.:{dept3}"
                ws["B9"].value = fmt_intlike(f_order)

                append_keycode_to_a10(ws, extract_inside_brackets(i_prod))
                ws["A11"].value = f"DESCRIPTION：{j_ename}" if j_ename else "DESCRIPTION："

                if gross_wt:
                    ws["E9"].value = f"GRS.WT.: {gross_wt} KGS"

                ws["E10"].value = f"D:{length_v}×{width_v}×{height_v}CMS"
                ws["C12"].value = carton_count

                out_name = f"{e_sales} {i_prod}#{h_cname} 唛头（TG）.xlsx"

            out_name = sanitize_filename(out_name)
            out_path = unique_path(out_dir, out_name)

            wb.save(out_path)
            success += 1
            print(f"✅ 第 {idx+2} 行完成 → {os.path.basename(out_path)}")

        except Exception as e:
            failed += 1
            print(f"❌ 第 {idx+2} 行失败：{e}")

    print(f"\n🎉 完成：成功 {success} 个，失败 {failed} 个。保存路径：{out_dir}")


if __name__ == "__main__":
    main()
